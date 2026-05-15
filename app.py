from flask import (Flask, render_template, request, jsonify,
                   send_file, session, redirect, url_for)
import os, json, random, time, sqlite3, io
from datetime import datetime
from functools import wraps
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import uuid
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'test_app_secret_2024_xyz_v6')
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False

BASE_DIR    = os.environ.get('RAILWAY_VOLUME_MOUNT_PATH',
              os.path.dirname(os.path.abspath(__file__)))
TESTS_DIR   = os.path.join(BASE_DIR, 'tests')
RESULTS_DIR = os.path.join(BASE_DIR, 'results')
CONFIG_FILE = os.path.join(BASE_DIR, 'config.json')
DB_FILE     = os.path.join(BASE_DIR, 'results.db')

DEFAULT_CONFIG = {"admin_password": generate_password_hash("admin123")}

# ═══════════════════════════════════════════
# DATABASE
# ═══════════════════════════════════════════
def get_db():
    con = sqlite3.connect(DB_FILE)
    con.row_factory = sqlite3.Row
    return con

def init_db():
    os.makedirs(TESTS_DIR,   exist_ok=True)
    os.makedirs(RESULTS_DIR, exist_ok=True)
    con = get_db()
    con.executescript('''
        CREATE TABLE IF NOT EXISTS results (
            id TEXT PRIMARY KEY, fio TEXT, grp TEXT, category TEXT,
            total INTEGER, correct INTEGER, wrong INTEGER,
            percentage REAL, date TEXT, answers TEXT, teacher_id TEXT);

        CREATE TABLE IF NOT EXISTS active_tests (
            token TEXT PRIMARY KEY, fio TEXT, grp TEXT, category TEXT,
            questions TEXT, start_time REAL, time_limit INTEGER,
            answers TEXT, created_at REAL, teacher_id TEXT);

        CREATE TABLE IF NOT EXISTS teachers (
            id TEXT PRIMARY KEY, name TEXT, username TEXT UNIQUE,
            password_hash TEXT, email TEXT, created_at TEXT,
            is_active INTEGER DEFAULT 1,
            time_limit INTEGER DEFAULT 30,
            question_count INTEGER DEFAULT 10);
    ''')
    con.commit(); con.close()

# ═══════════════════════════════════════════
# CONFIG  (admin paroli hash saqlanadi)
# ═══════════════════════════════════════════
def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE,'r',encoding='utf-8') as f:
            return json.load(f)
    return DEFAULT_CONFIG.copy()

def save_config(cfg):
    with open(CONFIG_FILE,'w',encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def check_admin_password(raw):
    cfg = load_config()
    h = cfg.get('admin_password','')
    # Eski tizim: oddiy matn parol (migratsiya uchun)
    if not h.startswith('pbkdf2') and not h.startswith('scrypt'):
        return h == raw
    return check_password_hash(h, raw)

def set_admin_password(raw):
    cfg = load_config()
    cfg['admin_password'] = generate_password_hash(raw)
    save_config(cfg)

# ═══════════════════════════════════════════
# ADMIN AUTH
# ═══════════════════════════════════════════
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated

# ═══════════════════════════════════════════
# TEACHER AUTH
# ═══════════════════════════════════════════
def teacher_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'teacher_id' not in session:
            return redirect(url_for('teacher_login'))
        return f(*args, **kwargs)
    return decorated

def get_current_teacher():
    tid = session.get('teacher_id')
    if not tid: return None
    con = get_db()
    row = con.execute('SELECT * FROM teachers WHERE id=?',(tid,)).fetchone()
    con.close()
    return dict(row) if row else None

# ═══════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════
def get_teacher_tests_dir(teacher_id):
    d = os.path.join(TESTS_DIR, f'teacher_{teacher_id}')
    os.makedirs(d, exist_ok=True)
    return d

def get_teacher_categories(teacher_id):
    d = get_teacher_tests_dir(teacher_id)
    cats = {}
    for fname in sorted(os.listdir(d)):
        if fname.endswith('.xlsx'):
            cats[fname[:-5]] = os.path.join(d, fname)
    return cats

def load_questions(filepath):
    df = pd.read_excel(filepath, header=None)
    questions = []
    for _, row in df.iterrows():
        vals = [str(v).strip() for v in row.tolist()
                if str(v).strip() not in ('nan','')]
        if len(vals) >= 2:
            questions.append({'question':vals[0],'correct':vals[1],'options':vals[1:]})
    return questions

def get_on_categories(categories):
    on1 = [k for k in categories if '1-' in k and ('ON' in k.upper() or 'ОН' in k)]
    on2 = [k for k in categories if '2-' in k and ('ON' in k.upper() or 'ОН' in k)]
    return on1, on2

def get_test_token():
    return request.cookies.get('test_token') or request.args.get('token')

def get_all_teachers(include_inactive=False):
    con = get_db()
    if include_inactive:
        rows = con.execute('SELECT * FROM teachers ORDER BY name').fetchall()
    else:
        rows = con.execute('SELECT * FROM teachers WHERE is_active=1 ORDER BY name').fetchall()
    con.close()
    return [dict(r) for r in rows]

def get_teacher_by_id(tid):
    con = get_db()
    row = con.execute('SELECT * FROM teachers WHERE id=?',(tid,)).fetchone()
    con.close()
    return dict(row) if row else None

# ═══════════════════════════════════════════
# DB RESULTS
# ═══════════════════════════════════════════
def db_save_result(r, teacher_id=None):
    con = get_db()
    con.execute('''INSERT OR REPLACE INTO results
        (id,fio,grp,category,total,correct,wrong,percentage,date,answers,teacher_id)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
        (r['id'],r['fio'],r['group'],r['category'],r['total'],r['correct'],
         r['wrong'],r['percentage'],r['date'],
         json.dumps(r['answers'],ensure_ascii=False), teacher_id))
    con.commit(); con.close()

def db_load_results(category=None, teacher_id=None):
    con = get_db()
    q = 'SELECT * FROM results WHERE 1=1'
    params = []
    if teacher_id:
        q += ' AND teacher_id=?'; params.append(teacher_id)
    if category and category != 'all':
        q += ' AND category=?'; params.append(category)
    q += ' ORDER BY date DESC'
    rows = con.execute(q, params).fetchall()
    con.close()
    out = []
    for row in rows:
        d = dict(row); d['group'] = d.pop('grp')
        d['answers'] = json.loads(d['answers']); out.append(d)
    return out

def save_active_test(token, data, teacher_id=None):
    con = get_db()
    con.execute('''INSERT OR REPLACE INTO active_tests
        (token,fio,grp,category,questions,start_time,time_limit,answers,created_at,teacher_id)
        VALUES (?,?,?,?,?,?,?,?,?,?)''',
        (token,data['fio'],data['group'],data['category'],
         json.dumps(data['questions'],ensure_ascii=False),
         data['start_time'],data['time_limit'],
         json.dumps(data['answers'],ensure_ascii=False),
         time.time(), teacher_id))
    con.commit(); con.close()

def load_active_test(token):
    con = get_db()
    row = con.execute('SELECT * FROM active_tests WHERE token=?',(token,)).fetchone()
    con.close()
    if not row: return None
    d = dict(row); d['group'] = d.pop('grp')
    d['questions'] = json.loads(d['questions'])
    d['answers']   = json.loads(d['answers'])
    return d

def delete_active_test(token):
    con = get_db()
    con.execute('DELETE FROM active_tests WHERE token=?',(token,))
    con.commit(); con.close()

def update_active_test_answers(token, answers):
    con = get_db()
    con.execute('UPDATE active_tests SET answers=? WHERE token=?',
                (json.dumps(answers,ensure_ascii=False), token))
    con.commit(); con.close()

# ═══════════════════════════════════════════
# TALABA ROUTES
# ═══════════════════════════════════════════
@app.route('/')
def index():
    teachers = get_all_teachers()
    return render_template('index.html', teachers=teachers)

@app.route('/api/teacher_categories/<teacher_id>')
def api_teacher_categories(teacher_id):
    cats = get_teacher_categories(teacher_id)
    on1, on2 = get_on_categories(cats)
    result = list(cats.keys())
    if on1 and on2: result.append('YaN')
    return jsonify({'categories': result})

@app.route('/start_test', methods=['POST'])
def start_test():
    data = request.json
    fio        = data.get('fio','').strip()
    group      = data.get('group','').strip()
    category   = data.get('category','')
    teacher_id = data.get('teacher_id','')
    if not fio or not group:
        return jsonify({'error':'FIO va guruh kiritilishi shart!'}),400
    if not teacher_id:
        return jsonify({'error':"O'qituvchini tanlang!"}),400
    if not category:
        return jsonify({'error':'Test turini tanlang!'}),400
    teacher = get_teacher_by_id(teacher_id)
    if not teacher:
        return jsonify({'error':"O'qituvchi topilmadi!"}),400
    cats = get_teacher_categories(teacher_id)
    questions = []
    q_count = teacher.get('question_count',10)
    t_limit = teacher.get('time_limit',30)
    if category == 'YaN':
        on1, on2 = get_on_categories(cats)
        if not on1 or not on2:
            return jsonify({'error':'YaN uchun 1-ON va 2-ON fayllari zarur!'}),400
        half = q_count//2
        q1 = load_questions(cats[on1[0]]); q2 = load_questions(cats[on2[0]])
        random.shuffle(q1); random.shuffle(q2)
        questions = q1[:half] + q2[q_count-half:]
    else:
        if category not in cats:
            return jsonify({'error':'Kategoriya topilmadi!'}),400
        all_q = load_questions(cats[category])
        random.shuffle(all_q)
        questions = all_q[:q_count]
    for q in questions:
        opts = q['options'][:]
        random.shuffle(opts)
        q['shuffled_options'] = opts
    token = str(uuid.uuid4())
    test_data = {
        'id':str(uuid.uuid4()),'fio':fio,'group':group,'category':category,
        'questions':questions,'start_time':time.time(),
        'time_limit':t_limit*60,'answers':[],'teacher_id':teacher_id
    }
    save_active_test(token, test_data, teacher_id)
    resp = jsonify({'success':True,'total':len(questions),'token':token})
    resp.set_cookie('test_token', token, max_age=7200, samesite='Lax')
    return resp

@app.route('/test')
def test_page():
    token = get_test_token()
    if not token:
        return render_template('error.html', msg='Test topilmadi. Qaytadan boshlang.')
    t = load_active_test(token)
    if not t:
        return render_template('error.html', msg='Test topilmadi. Qaytadan boshlang.')
    remaining = max(0, t['time_limit']-(time.time()-t['start_time']))
    teacher = get_teacher_by_id(t.get('teacher_id','')) or {}
    return render_template('test.html',
                           fio=t['fio'], group=t['group'], category=t['category'],
                           total=len(t['questions']), time_limit=int(remaining),
                           token=token, teacher_name=teacher.get('name',''))

@app.route('/get_question/<int:idx>')
def get_question(idx):
    token = get_test_token()
    if not token: return jsonify({'error':'Token topilmadi'}),400
    t = load_active_test(token)
    if not t: return jsonify({'error':'Test topilmadi'}),400
    elapsed = time.time()-t['start_time']
    if elapsed > t['time_limit']: return jsonify({'timeout':True})
    if idx >= len(t['questions']): return jsonify({'done':True})
    q = t['questions'][idx]
    return jsonify({'question':q['question'],'options':q['shuffled_options'],
                    'index':idx,'total':len(t['questions']),'remaining':int(t['time_limit']-elapsed)})

@app.route('/submit_answer', methods=['POST'])
def submit_answer():
    token = get_test_token()
    if not token: return jsonify({'error':'Token topilmadi'}),400
    t = load_active_test(token)
    if not t: return jsonify({'error':'Test topilmadi'}),400
    data = request.json
    q = t['questions'][data['index']]
    correct = (data['answer'] == q['correct'])
    t['answers'].append({'index':data['index'],'answer':data['answer'],'correct':correct})
    update_active_test_answers(token, t['answers'])
    return jsonify({'correct':correct,'correct_answer':q['correct']})

@app.route('/finish_test', methods=['POST'])
def finish_test():
    token = get_test_token()
    if not token: return jsonify({'error':'Token topilmadi'}),400
    t = load_active_test(token)
    if not t: return jsonify({'error':'Test topilmadi'}),400
    total = len(t['questions'])
    correct_count = sum(1 for a in t['answers'] if a['correct'])
    percentage = round(correct_count/total*100,1) if total>0 else 0
    result = {
        'id':str(uuid.uuid4()),'fio':t['fio'],'group':t['group'],'category':t['category'],
        'total':total,'correct':correct_count,'wrong':total-correct_count,
        'percentage':percentage,'date':datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'answers':t['answers']
    }
    db_save_result(result, t.get('teacher_id'))
    delete_active_test(token)
    resp = jsonify(result)
    resp.delete_cookie('test_token')
    return resp

# ═══════════════════════════════════════════
# O'QITUVCHI ROUTES
# ═══════════════════════════════════════════
@app.route('/teacher/login', methods=['GET','POST'])
def teacher_login():
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        con = get_db()
        t = con.execute('SELECT * FROM teachers WHERE username=? AND is_active=1',(username,)).fetchone()
        con.close()
        if t and check_password_hash(t['password_hash'], password):
            session['teacher_id']   = t['id']
            session['teacher_name'] = t['name']
            return redirect(url_for('teacher_dashboard'))
        return render_template('teacher_login.html', error="Login yoki parol noto'g'ri!")
    return render_template('teacher_login.html', error=None)

@app.route('/teacher/register', methods=['GET','POST'])
def teacher_register():
    if request.method == 'POST':
        name      = request.form.get('name','').strip()
        username  = request.form.get('username','').strip()
        email     = request.form.get('email','').strip()
        password  = request.form.get('password','')
        password2 = request.form.get('password2','')
        if not all([name, username, password]):
            return render_template('teacher_register.html', error="Barcha majburiy maydonlarni to'ldiring!")
        if password != password2:
            return render_template('teacher_register.html', error='Parollar mos kelmadi!')
        if len(password) < 6:
            return render_template('teacher_register.html', error='Parol kamida 6 ta belgi!')
        con = get_db()
        ex = con.execute('SELECT id FROM teachers WHERE username=?',(username,)).fetchone()
        if ex:
            con.close()
            return render_template('teacher_register.html', error='Bu login band!')
        tid = str(uuid.uuid4())
        con.execute('''INSERT INTO teachers
            (id,name,username,password_hash,email,created_at,time_limit,question_count)
            VALUES (?,?,?,?,?,?,?,?)''',
            (tid,name,username,generate_password_hash(password),email,
             datetime.now().strftime('%Y-%m-%d %H:%M:%S'),30,10))
        con.commit(); con.close()
        session['teacher_id']   = tid
        session['teacher_name'] = name
        return redirect(url_for('teacher_dashboard'))
    return render_template('teacher_register.html', error=None)

@app.route('/teacher/logout')
def teacher_logout():
    session.pop('teacher_id',None); session.pop('teacher_name',None)
    return redirect(url_for('teacher_login'))

@app.route('/teacher')
@teacher_login_required
def teacher_dashboard():
    teacher = get_current_teacher()
    t_dir   = get_teacher_tests_dir(teacher['id'])
    my_cats = [f[:-5] for f in sorted(os.listdir(t_dir)) if f.endswith('.xlsx')]
    results = db_load_results(teacher_id=teacher['id'])
    all_cats = list(set([r['category'] for r in results]))
    return render_template('teacher_dashboard.html',
                           teacher=teacher, my_categories=my_cats,
                           result_categories=all_cats, results=results)

@app.route('/teacher/save_settings', methods=['POST'])
@teacher_login_required
def teacher_save_settings():
    teacher = get_current_teacher()
    data = request.json
    con = get_db()
    con.execute('UPDATE teachers SET time_limit=?, question_count=? WHERE id=?',
                (int(data.get('time_limit',30)), int(data.get('question_count',10)), teacher['id']))
    con.commit(); con.close()
    return jsonify({'success':True})

@app.route('/teacher/upload_test', methods=['POST'])
@teacher_login_required
def teacher_upload_test():
    teacher = get_current_teacher()
    if 'file' not in request.files: return jsonify({'error':'Fayl yuklanmadi'}),400
    f    = request.files['file']
    name = request.form.get('name','').strip()
    if not name: return jsonify({'error':'Test nomi kiritilmagan'}),400
    if not f.filename.endswith('.xlsx'): return jsonify({'error':'Faqat xlsx fayl'}),400
    t_dir = get_teacher_tests_dir(teacher['id'])
    f.save(os.path.join(t_dir, f"{name}.xlsx"))
    return jsonify({'success':True,'message':f'"{name}" muvaffaqiyatli yuklandi!'})

@app.route('/teacher/delete_test', methods=['POST'])
@teacher_login_required
def teacher_delete_test():
    teacher = get_current_teacher()
    name  = request.json.get('name','')
    t_dir = get_teacher_tests_dir(teacher['id'])
    path  = os.path.join(t_dir, f"{name}.xlsx")
    if os.path.exists(path): os.remove(path); return jsonify({'success':True})
    return jsonify({'error':'Fayl topilmadi'}),404

@app.route('/teacher/export_results', methods=['POST'])
@teacher_login_required
def teacher_export_results():
    teacher = get_current_teacher()
    data = request.json
    cat  = data.get('category','all')
    results = db_load_results(category=cat if cat!='all' else None, teacher_id=teacher['id'])
    if not results: return jsonify({'error':'Natijalar topilmadi'}),404
    return _build_excel_results(results, cat)

# ═══════════════════════════════════════════
# ADMIN ROUTES
# ═══════════════════════════════════════════
@app.route('/admin/login', methods=['GET','POST'])
def admin_login():
    if session.get('admin_logged_in'):
        return redirect(url_for('admin_dashboard'))
    error = None
    if request.method == 'POST':
        pw = request.form.get('password','')
        if check_admin_password(pw):
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        error = "Parol noto'g'ri!"
    return render_template('admin_login.html', error=error)

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

@app.route('/admin')
@app.route('/admin/')
@admin_required
def admin_dashboard():
    teachers = get_all_teachers(include_inactive=True)
    # Har bir o'qituvchi uchun statistika
    con = get_db()
    stats = []
    for t in teachers:
        t_dir  = get_teacher_tests_dir(t['id'])
        tests  = [f[:-5] for f in os.listdir(t_dir) if f.endswith('.xlsx')]
        # Talabalar natijalari
        rows = con.execute(
            'SELECT percentage FROM results WHERE teacher_id=?',(t['id'],)).fetchall()
        done    = len(rows)
        passed  = sum(1 for r in rows if r['percentage'] >= 55)
        failed  = done - passed
        stats.append({
            'teacher': t,
            'tests':   tests,
            'tests_count': len(tests),
            'done':    done,
            'passed':  passed,
            'failed':  failed,
        })
    con.close()
    return render_template('admin.html', stats=stats)

@app.route('/admin/change_password', methods=['POST'])
@admin_required
def admin_change_password():
    data    = request.json
    cur_pw  = data.get('current_password','')
    new_pw  = data.get('new_password','')
    if not check_admin_password(cur_pw):
        return jsonify({'error':"Joriy parol noto'g'ri!"}),403
    if len(new_pw) < 6:
        return jsonify({'error':'Yangi parol kamida 6 ta belgi!'}),400
    set_admin_password(new_pw)
    return jsonify({'success':True})

# Foydalanuvchi (teacher) CRUD
@app.route('/admin/teacher/<tid>')
@admin_required
def admin_teacher_profile(tid):
    t = get_teacher_by_id(tid)
    if not t: return render_template('error.html', msg="O'qituvchi topilmadi"), 404
    t_dir  = get_teacher_tests_dir(tid)
    tests  = [f[:-5] for f in sorted(os.listdir(t_dir)) if f.endswith('.xlsx')]
    results = db_load_results(teacher_id=tid)
    return render_template('admin_teacher.html', teacher=t, tests=tests, results=results)

@app.route('/admin/teacher/add', methods=['POST'])
@admin_required
def admin_add_teacher():
    data     = request.json
    name     = data.get('name','').strip()
    username = data.get('username','').strip()
    email    = data.get('email','').strip()
    password = data.get('password','')
    if not all([name, username, password]):
        return jsonify({'error':"Ism, login va parol majburiy!"}),400
    if len(password) < 6:
        return jsonify({'error':'Parol kamida 6 ta belgi!'}),400
    con = get_db()
    ex = con.execute('SELECT id FROM teachers WHERE username=?',(username,)).fetchone()
    if ex:
        con.close()
        return jsonify({'error':'Bu login allaqachon band!'}),400
    tid = str(uuid.uuid4())
    con.execute('''INSERT INTO teachers
        (id,name,username,password_hash,email,created_at,time_limit,question_count)
        VALUES (?,?,?,?,?,?,?,?)''',
        (tid,name,username,generate_password_hash(password),email,
         datetime.now().strftime('%Y-%m-%d %H:%M:%S'),30,10))
    con.commit(); con.close()
    return jsonify({'success':True, 'id':tid})

@app.route('/admin/teacher/edit', methods=['POST'])
@admin_required
def admin_edit_teacher():
    data     = request.json
    tid      = data.get('id','')
    name     = data.get('name','').strip()
    username = data.get('username','').strip()
    email    = data.get('email','').strip()
    password = data.get('password','').strip()
    tl       = int(data.get('time_limit',30))
    qc       = int(data.get('question_count',10))
    con = get_db()
    if password:
        if len(password) < 6:
            con.close()
            return jsonify({'error':'Parol kamida 6 ta belgi!'}),400
        con.execute('''UPDATE teachers SET name=?,username=?,email=?,
            password_hash=?,time_limit=?,question_count=? WHERE id=?''',
            (name,username,email,generate_password_hash(password),tl,qc,tid))
    else:
        con.execute('''UPDATE teachers SET name=?,username=?,email=?,
            time_limit=?,question_count=? WHERE id=?''',
            (name,username,email,tl,qc,tid))
    con.commit(); con.close()
    return jsonify({'success':True})

@app.route('/admin/teacher/toggle', methods=['POST'])
@admin_required
def admin_toggle_teacher():
    data   = request.json
    tid    = data.get('id','')
    active = data.get('active',1)
    con = get_db()
    con.execute('UPDATE teachers SET is_active=? WHERE id=?',(active,tid))
    con.commit(); con.close()
    return jsonify({'success':True})

@app.route('/admin/teacher/delete', methods=['POST'])
@admin_required
def admin_delete_teacher():
    data = request.json
    tid  = data.get('id','')
    con = get_db()
    con.execute('DELETE FROM teachers WHERE id=?',(tid,))
    con.commit(); con.close()
    return jsonify({'success':True})

@app.route('/admin/teacher/reset_password', methods=['POST'])
@admin_required
def admin_reset_password():
    data = request.json
    tid  = data.get('id','')
    new_pw = data.get('password','')
    if len(new_pw) < 6:
        return jsonify({'error':'Parol kamida 6 ta belgi!'}),400
    con = get_db()
    con.execute('UPDATE teachers SET password_hash=? WHERE id=?',
                (generate_password_hash(new_pw), tid))
    con.commit(); con.close()
    return jsonify({'success':True})

@app.route('/teacher/export_results_by_id/<tid>', methods=['POST'])
@admin_required
def admin_export_teacher_results(tid):
    results = db_load_results(teacher_id=tid)
    if not results: return jsonify({'error':'Natijalar topilmadi'}),404
    return _build_excel_results(results, f'teacher_{tid[:8]}')

@app.route('/admin/export_report', methods=['GET'])
@admin_required
def admin_export_report():
    """Barcha o'qituvchilar bo'yicha umumiy hisobot xlsx"""
    teachers = get_all_teachers(include_inactive=True)
    con = get_db()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Hisobot"

    h_fill  = PatternFill("solid", fgColor="1A3A6B")
    h_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    border  = Border(left=Side(style='thin'),right=Side(style='thin'),
                     top=Side(style='thin'),bottom=Side(style='thin'))
    e_fill  = PatternFill("solid", fgColor="EBF5FB")

    headers = ["№", "Foydalanuvchi (F.I.O.)", "Login", "Fan / Test nomi",
               "Testlar soni", "Ishlanganlar soni",
               "O'tganlar (≥55%)", "O'tmaganlar", "O'tish foizi (%)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = h_fill; cell.font = h_font
        cell.alignment = Alignment(horizontal='center',vertical='center')
        cell.border = border
    ws.row_dimensions[1].height = 28

    row_num = 2
    for i, t in enumerate(teachers, 1):
        t_dir  = get_teacher_tests_dir(t['id'])
        tests  = [f[:-5] for f in os.listdir(t_dir) if f.endswith('.xlsx')]
        rows_r = con.execute(
            'SELECT percentage FROM results WHERE teacher_id=?',(t['id'],)).fetchall()
        done   = len(rows_r)
        passed = sum(1 for r in rows_r if r['percentage'] >= 55)
        failed = done - passed
        pct    = round(passed/done*100,1) if done>0 else 0

        fan_nomi = ', '.join(tests) if tests else '—'
        ws.append([i, t['name'], t['username'], fan_nomi,
                   len(tests), done, passed, failed, pct])

        for cell in ws[row_num]:
            cell.border = border
            cell.alignment = Alignment(horizontal='center',vertical='center')
            if i % 2 == 0: cell.fill = e_fill
        # O'tish foizi rangi
        pct_cell = ws.cell(row=row_num, column=9)
        if isinstance(pct_cell.value, (int,float)):
            color = "1E8449" if pct_cell.value>=80 else ("D4AC0D" if pct_cell.value>=55 else "C0392B")
            pct_cell.font = Font(color=color, bold=True, name="Arial")
        row_num += 1

    col_widths = [5, 28, 18, 32, 14, 18, 16, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    con.close()
    os.makedirs(RESULTS_DIR, exist_ok=True)
    fname = f"admin_hisobot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fpath = os.path.join(RESULTS_DIR, fname)
    wb.save(fpath)
    return send_file(fpath, as_attachment=True, download_name=fname)

# ═══════════════════════════════════════════
# EXCEL RESULTS helper
# ═══════════════════════════════════════════
def _build_excel_results(results, label):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="Natijalar"
    h_fill = PatternFill("solid",fgColor="1A3A6B")
    h_font = Font(bold=True,color="FFFFFF",name="Arial",size=11)
    border = Border(left=Side(style='thin'),right=Side(style='thin'),
                    top=Side(style='thin'),bottom=Side(style='thin'))
    e_fill = PatternFill("solid",fgColor="EBF5FB")
    ws.append(["№","F.I.O.","Guruh","Kategoriya","Jami","To'g'ri","Xato","Foiz (%)","Sana"])
    for cell in ws[1]:
        cell.fill=h_fill; cell.font=h_font
        cell.alignment=Alignment(horizontal='center',vertical='center')
        cell.border=border
    ws.row_dimensions[1].height = 28
    for i,r in enumerate(results,1):
        ws.append([i,r['fio'],r['group'],r['category'],r['total'],
                   r['correct'],r['wrong'],r['percentage'],r['date']])
        for cell in ws[ws.max_row]:
            cell.border=border
            cell.alignment=Alignment(horizontal='center',vertical='center')
            if i%2==0: cell.fill=e_fill
        pct = ws.cell(row=ws.max_row,column=8)
        if isinstance(pct.value,(int,float)):
            color="1E8449" if pct.value>=85 else ("D4AC0D" if pct.value>=55 else "C0392B")
            pct.font=Font(color=color,bold=True,name="Arial")
    for i,w in enumerate([5,30,15,20,12,12,10,12,22],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    os.makedirs(RESULTS_DIR,exist_ok=True)
    fname = f"natijalar_{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fpath = os.path.join(RESULTS_DIR,fname)
    wb.save(fpath)
    return send_file(fpath,as_attachment=True,download_name=fname)

# ═══════════════════════════════════════════
# INIT
# ═══════════════════════════════════════════
with app.app_context():
    init_db()
    if not os.path.exists(CONFIG_FILE): save_config(DEFAULT_CONFIG)

if __name__=='__main__':
    port = int(os.environ.get('PORT',5050))
    app.run(host='0.0.0.0', port=port, debug=False)
